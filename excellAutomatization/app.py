import openpyxl

# Open the source Excel file
source_file_path = "test_1.xlsx"
source_workbook = openpyxl.load_workbook(source_file_path)
source_sheet = source_workbook.active

# Create a new workbook
new_workbook = openpyxl.Workbook()
new_sheet = new_workbook.active

# Copy the column name separately
column_name = source_sheet.cell(row=1, column=1).value
new_sheet.cell(row=1, column=1, value=column_name)

# Iterate through the remaining rows in the source sheet and copy the first column
for row in source_sheet.iter_rows(min_row=2, values_only=True):
    new_sheet.cell(row=new_sheet.max_row + 1, column=1, value=row[0])

# Save the new workbook
new_file_path = "test_2.xlsx"
new_workbook.save(new_file_path)

# Close the workbooks
source_workbook.close()
new_workbook.close()

print("Column copied and pasted successfully!")
